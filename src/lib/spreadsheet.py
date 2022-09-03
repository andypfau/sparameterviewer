from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Color


def clean_name(name: str) -> str:
    result = ''
    for c in name:
        if c in 'abcdefghijklmnopqrstuvwxyz_ABCDEFGHIJKLMNOPQRSTUVWXYZ_0123456789_+-~#()':
            result += c
        elif c in '[{<':
            result += '('
        elif c in ']}>':
            result += ')'
        else:
            result += '_'
    return result


class SpreadsheetGen:

    def __init__(self):
        self.sheets = []
        self.wb = Workbook()


    def add_sheet(self, name):
        if len(self.sheets)==0:
            ws = self.wb.active
            ws.title = clean_name(name)
        else:
            ws = self.wb.create_sheet(clean_name(name))
        self.sheets.append(SpreadsheetGen.Sheet(self, ws))
        return self.sheets[-1]


    def save(self, filename):
        self.wb.save(filename)


    class Sheet:

        def __init__(self, parent, ws):
            self.parent = parent
            self.ws = ws
            self.row = 0
            self.col = 0


        def col_widths(self, widths):
            for i,w in enumerate(widths):
                self.ws.column_dimensions[get_column_letter(i+1)].width = w


        def _coord(self, i_row,i_col):
            return get_column_letter(i_col+1)+str(i_row+1)


        def cell(self, row, col, value, font=None):
            self.ws[self._coord(row,col)] = value
            if font:
                self.ws[self._coord(row,col)].font = font


        def add_row(self, cols, font=None):
            for i,col in enumerate(cols):
                self.cell(self.row, i, col, font)
            self.row += 1


        def add_table(self, header_row=None, rows=None, style=None, name=None, header_font=None, font=None):
            styled = (style is not None) or (name is not None)
            first_row = self.row
            if header_row:
                self.add_row(header_row, header_font)
            for row in rows:
                self.add_row(row, font)
            if styled:
                if not header_row:
                    raise Exception('A table must have a header row')
                if (style is None) or (name is None):
                    raise Exception('A table must have both a name and a style')
                n_cols = max(len(header_row), max([len(row) for row in rows]))
                self.ws.add_table(Table(
                    ref=self._coord(first_row,0)+':'+self._coord(self.row-1,n_cols-1),
                    displayName=clean_name(name),
                    tableStyleInfo=style
                ))
